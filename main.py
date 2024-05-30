from tkinter import *
from PIL import Image,ImageTk

from tkcalendar import DateEntry
from tkinter import messagebox



# Create object
root = Tk()
root.title("vignan institute of infromation technology college portal",)

# Adjust size
w,h=root.winfo_screenwidth(),root.winfo_screenheight()
root.geometry('%dx%d+0+0' % (w,h))


def details():
    from tkinter import messagebox
    import pandas as pd

    def submit():
        data = {
            "Student Name": [ent1.get()],
            "Registration Number": [ent2.get()],
            "Course": [ent3.get()],
            "Branch": [ent4.get()],
            "Email Address": [ent5.get()],
            "Mobile Number": [ent6.get()],
            "State": [ent7.get()],
            "District": [ent8.get()],
            "Gender": [ent9.get()],
            "Seat Type": [ent10.get()]
        }
        new_row_df = pd.DataFrame(data)



        # Check if the file already exists
        if os.path.exists("database.xlsx"):
            # Read the existing Excel file
            existing_df = pd.read_excel("database.xlsx")
            # Concatenate the existing DataFrame with the new row DataFrame
            updated_df = pd.concat([existing_df, new_row_df], ignore_index=True)
            # Save the updated DataFrame to the Excel file without overwriting the existing data
            updated_df.to_excel("database.xlsx", index=False)
        else:
            # If the file doesn't exist, save the new data to a new Excel file
            new_row_df.to_excel("database.xlsx", index=False)

        messagebox.showinfo("Confirmation", "Your application has been successfully submitted")

    root = Tk()
    root.geometry("850x650")
    root.title("Student Application Form for Attendance")
    root.config(bg="honeydew2")

    Label(root, text="Fill the student details here", font=("bold",17), background="honeydew2").place(x=290, y=20)

    name_var = StringVar()
    reg_var = StringVar()
    cour_var = StringVar()
    branch_var = StringVar()
    email_var = StringVar()
    mobile_var = StringVar()
    state_var = StringVar()
    dist_var = StringVar()
    gender_var = StringVar()
    seat_var = StringVar()

    Label(root, text="Enter student name  :", font=("calibre",10), background="honeydew2").place(x=30, y=60)
    Label(root, text="Enter registration number:", font=("calibre",10), background="honeydew2").place(x=30, y=100)
    Label(root, text="Enter course      :", font=("calibre",10), background="honeydew2").place(x=30, y=140)
    Label(root, text="Enter branch    :", font=("calibre",10), background="honeydew2").place(x=30, y=180)
    Label(root, text="Enter email address     :", font=("calibre",10), background="honeydew2").place(x=30, y=220)

    Label(root, text="Enter mobile number     :", font=("calibre",10), background="honeydew2").place(x=450, y=60)
    Label(root, text="Enter the state           :", font=("calibre",10), background="honeydew2").place(x=450, y=100)
    Label(root, text="Enter the district         :", font=("calibre",10), background="honeydew2").place(x=450, y=140)
    Label(root, text="Enter gender       :", font=("calibre",10), background="honeydew2").place(x=450, y=180)
    Label(root, text="Enter seat type(jvd)      :", font=("calibre",10), background="honeydew2").place(x=450, y=220)

    ent1 = Entry(root, width=30, textvariable=name_var)
    ent1.place(x=220, y=60)
    ent2 = Entry(root, width=30, textvariable=reg_var)
    ent2.place(x=220, y=100)
    ent3 = Entry(root, width=30, textvariable=cour_var)
    ent3.place(x=220, y=140)
    ent4 = Entry(root, width=30, textvariable=branch_var)
    ent4.place(x=220, y=180)
    ent5 = Entry(root, width=30, textvariable=email_var)
    ent5.place(x=220, y=220)
    ent6 = Entry(root, width=30, textvariable=mobile_var)
    ent6.place(x=640, y=60)
    ent7 = Entry(root, width=30, textvariable=state_var)
    ent7.place(x=640, y=100)
    ent8 = Entry(root, width=30, textvariable=dist_var)
    ent8.place(x=640, y=140)
    ent9 = Entry(root, width=30, textvariable=gender_var)
    ent9.place(x=640, y=180)
    ent10 = Entry(root, width=30, textvariable=seat_var)
    ent10.place(x=640, y=220)

    button1 = Button(root, text="Submit", width=10, bg="black", foreground="white", font=("calibre",12), command=submit)
    button1.place(x=400, y=330, height=20)

    root.mainloop()


def datab():

    root2 = Tk()
    root2.geometry("650x350")
    root2.title("data base of the students")
    root2.config(bg="honeydew2")


    def det():
        file_path = r"C:\Users\revan\PycharmProjects\face\database.xlsx"
        os.startfile(file_path)


    def images():
        file_path = r"C:\Users\revan\PycharmProjects\face\captured_images"
        os.startfile(file_path)

    b1 = Button(root2,text="student information",bg="black",fg="white",height=5,width=15,command=det)
    b1.place(x=100,y=100)


    b2 = Button(root2,text="student faces",bg="black",fg="white",height=5,width=15,command=images)
    b2.place(x=400,y=100)



########################################################         FOR   LOADING IMAGES IN CAPTURE FOLDER         ############################################
def capture():
    import cv2
    import os
    import tkinter as tk
    from tkinter import filedialog

    def capture_images(folder_path, folder_name):
        # Create the folder if it doesn't exist
        folder_path = os.path.join(folder_path, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Initialize webcam
        cap = cv2.VideoCapture(0)

        count = 0
        while count < 100:
            ret, frame = cap.read()
            if not ret:
                print("Failed to capture image")
                break

            # Display the captured frame
            cv2.imshow('Capture Image', frame)

            # Save the captured image
            img_name = os.path.join(folder_path, f"image_{count}.jpg")
            cv2.imwrite(img_name, frame)

            count += 1

            # Press 'q' to exit
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        # Release the webcam and close OpenCV windows
        cap.release()
        cv2.destroyAllWindows()

    def get_folder_and_name():
        folder_path = filedialog.askdirectory()
        if folder_path:
            folder_name = tk.simpledialog.askstring("Folder Name", "Enter folder name:")
            if folder_name:
                capture_images(folder_path, folder_name)

    # Create Tkinter window
    window = tk.Tk()
    window.title("Face Recognition Attendance")

    # Create button to select folder and enter folder name
    folder_button = tk.Button(window, text="Select Folder", command=get_folder_and_name)
    folder_button.pack(pady=20)

    window.mainloop()


##################################################################   TAKE ATTANDANCE USING CAPTUED IMAGES FOLDER     ########################33

def attendance():
    import os
    import cv2
    import datetime
    from openpyxl import Workbook, load_workbook

    def load_student_images(root_folder):
        student_images = {}

        for student_folder in os.listdir(root_folder):
            student_folder_path = os.path.join(root_folder, student_folder)
            if os.path.isdir(student_folder_path):
                student_name = student_folder
                student_images[student_name] = []
                for student_image_name in os.listdir(student_folder_path):
                    student_image_path = os.path.join(student_folder_path, student_image_name)
                    if student_image_path.endswith('.jpg') or student_image_path.endswith('.png'):
                        student_images[student_name].append(cv2.imread(student_image_path))
        return student_images

    def recognize_face(frame, student_images):
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        for (x, y, w, h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            for student_name, images in student_images.items():
                for stored_image in images:
                    # Compare the detected face with stored images
                    # Here, you can use any face recognition algorithm or method
                    if is_match(roi_gray, stored_image):
                        return student_name
        return None

    def is_match(roi_gray, stored_image):
        # Implement your face matching logic here
        # For example, you can use OpenCV's face recognition methods or any other library
        # For simplicity, let's assume a simple comparison method (e.g., histogram comparison)
        # This is just a placeholder, you should replace it with a proper face recognition algorithm
        return True  # Placeholder, replace with actual face matching logic

    def mark_attendance(student_name, attendance_sheet_path, marked_students):
        try:
            if student_name not in marked_students:
                if not os.path.exists(attendance_sheet_path):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['Student Name', 'Time'])
                else:
                    wb = load_workbook(attendance_sheet_path)
                    ws = wb.active
                ws.append([student_name, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                wb.save(attendance_sheet_path)
                marked_students.add(student_name)
                print("Attendance marked for", student_name)


        except Exception as e:
            print("Error occurred while marking attendance:", e)

    def main():
        root_folder = "captured_images"
        student_images = load_student_images(root_folder)
        attendance_sheet_path = "updated_attendance.xlsx"
        marked_students = set()

        cap = cv2.VideoCapture(0)

        while True:
            ret, frame = cap.read()
            if not ret:
                print("Failed to capture image")
                break

            student_name = recognize_face(frame, student_images)
            if student_name:
                mark_attendance(student_name, attendance_sheet_path, marked_students)
                cv2.putText(frame, "Student: " + student_name, (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

            cv2.imshow('Face Recognition', frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

    if __name__ == "__main__":
        main()



######################################################   TO OPEN EXCEL SHEET      #############################################33333

import os

def open_excel():
    file_path = r"C:\Users\revan\PycharmProjects\face\updated_attendance.xlsx"
    os.startfile(file_path)


#########################         BACK GROUND IMAGE     #####################################
b= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\bg.png")

bg=b.resize((w,h-2))

firstphoto = ImageTk.PhotoImage(bg)

firstbut = Label(root, image=firstphoto, borderwidth=5)
firstbut.place(x=-5, y=0)


#############################      LOGO IMAGE           #######################################
logo= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\Logo.jpg")
logo=logo.resize((1520,100))

logophoto = ImageTk.PhotoImage(logo)

logophotos = Label(root, image=logophoto, borderwidth=5)
logophotos.place(x=0, y=10)


#############################      STUDENTS DETAILS          #######################################
det= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\details.jpg")
det=det.resize((200,200))

detphoto = ImageTk.PhotoImage(det)

detphotos = Button(root, image=detphoto, borderwidth=0,command=details)
detphotos.place(x=220, y=170)



###################################    DATA BASE    ####################################
data= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\database3.jpg")
data=data.resize((200,200))

dataphoto = ImageTk.PhotoImage(data)

dataphotos = Button(root, image=dataphoto, borderwidth=0,command=datab)
dataphotos.place(x=640, y=170)




###########################  TO OPEN CAMERA   ###################################################

op= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\open _cam.jpg")
op=op.resize((200,200))

opphoto = ImageTk.PhotoImage(op)

opphotos = Button(root, image=opphoto, borderwidth=0,command=capture)
opphotos.place(x=1040, y=170)


###################################    TO TAKE ATTENDACE   ####################################
att= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\scan.jpg")
att=att.resize((200,200))

attphoto = ImageTk.PhotoImage(att)

attphotos = Button(root, image=attphoto, borderwidth=0,command=attendance)
attphotos.place(x=220, y=500)


###################################    TO OPEN EXCEL SHEET  ####################################
ex= Image.open("C:\\Users\\revan\\OneDrive\\Pictures\\sheet.png")
ex=ex.resize((200,200))

exphoto = ImageTk.PhotoImage(ex)

exphotos = Button(root, image=exphoto, borderwidth=0,command=open_excel)
exphotos.place(x=640, y=500)









######  student detail name   ####

detname = Button(text="student details",bg="black",fg="white",height=1,width=15,command=details)
detname.place(x=260,y=380)

######  data base name   ####

detname = Button(text="student database",bg="black",fg="white",height=1,width=15,command=datab)
detname.place(x=680,y=380)



######  open camera name   ####

detname = Button(text="load faces ",bg="black",fg="white",height=1,width=15,command=capture)
detname.place(x=1080,y=380)



######  take attendace   ####

detname = Button(text="take attendace ",bg="black",fg="white",height=1,width=15,command=attendance)
detname.place(x=260,y=720)



######  excel sheet   ####

detname = Button(text="excel sheet ",bg="black",fg="white",height=1,width=15,command=open_excel)
detname.place(x=680,y=720)


# For scrolling

def marquee_fun(widget,widget_w,widget_h,total_w,total_h,direction,speed,position=0):
    if direction=='right':
        if position>=total_w-widget_w:
            position=0
        position=position+speed
        widget.place(x=position)
    elif direction=='left':
        if position<0:
            position=total_w-widget_w
        position=position-speed
        widget.place(x=position)
    widget.after(10,lambda:marquee_fun(widget,widget_w,widget_h,total_w,total_h,direction,speed,position))


x=Label(root,text="Vignan's Institute of Information Technology (VIIT) was established in 2002 in the City of Destiny.",font=("times to roman",15),bg="gray4",fg="white")
x.place(x=0,y=120,width=900,height=30)


x.after(100,lambda:marquee_fun(x,150,30,1650,0,'left',1))




# Execute tkinter
root.mainloop()





